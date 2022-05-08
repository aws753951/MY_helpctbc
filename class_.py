from openpyxl import Workbook, load_workbook

import re
import pprint
import sys
import getopt

yearinput = 2021
monthinput = 5

short_opts = 'hy:m:'
long_opts = 'help year= month= '.split()
try:
    opts, args = getopt.getopt(sys.argv[1:], short_opts, long_opts)
except getopt.GetoptError:
    print('error, call 朱大師')
    sys.exit(2)
for opt, arg in opts:
    if opt in ("-y", "--year"):
        yearinput = arg
    elif opt in ("-m", "--month"):
        monthinput = arg



wb = load_workbook('2020_Classroom滿意度.xlsx')

titles = wb.sheetnames

dic = {}

for title in titles:
    ws = wb[title]
    i = 1
    dic[title] = []
    while True:
        i += 1
        time = ws['A' + str(i)].value
        if time == None:
            break
        year = re.search(r'(\d+)-(\d+)-(\d+)', str(time)).group(1)
        month = re.search(r'(\d+)-(\d+)-(\d+)', str(time)).group(2)
        if (int(year) == int(yearinput)) and (int(month) == int(monthinput)):
            dic[title].append([ws['B' + str(i)].value, ws['D' + str(i)].value])

name_teachers = dic.keys()
dic_grade = []
for teacher in name_teachers:
    if dic[teacher] == []:
        dic_grade.append([teacher, 0, 70])
        continue
    if teacher == '學員信箱名單':
        continue
    best = 0
    normal = 0
    bad = 0
    count_all = 0
    for comment in dic[teacher]:
        if comment[1] == '非常好':
            best += 1
        elif comment[1] == '好':
            normal += 1
        elif comment[1] == '普通':
            bad += 1
    count_all = best + normal + bad
    grade = 100*best/count_all + 85*normal/count_all + 70*bad/count_all
    dic_grade.append([teacher, count_all, grade])

biggest = 0
smallest = 100
for i in dic_grade:
    print(i[1])
    if (biggest == 0) or (i[1] > biggest): 
        biggest = i[1]
    if i[1] < smallest:
        smallest = i[1]

dis = biggest - smallest
for i in dic_grade:
    i[1] = i[1]/dis*30+70
pprint.pprint(dic_grade) # here1


members = {}

for teacher in name_teachers:
    if (dic[teacher] == []) or (dic[teacher] == '學員信箱名單'):
        continue
    raws = dic[teacher]
    for raw in raws:
        if raw[0] not in members:
            members[raw[0]] = 1
        else:
            members[raw[0]] += 1
pprint.pprint(members)   # here2

wb_output = Workbook()
ws_output = wb_output.active
ws_output.title = '分析師成績'
wb_output.create_sheet('區域成績')

ws_output.append(['分析師', '該月人氣度', '該月滿意度'])
for i in dic_grade:
    ws_output.append(i)
ws_output2 = wb_output['區域成績']
c = []
for i in members:
    c.append([i, members[i]])
ws_output2.append(['學員', '該月觀看次數'])
for i in c:
    ws_output2.append(i)

wb_output.save(f'classroom成績{yearinput}-{monthinput}.xlsx')
    

# ws = wb['Form Responses 86']



# for row in range(2, 29):
#     tim = ws['A' + str(row)].value
#     re.search(r'(\d+)-(\d+)-(\d+)', str(tim)).group(2)
